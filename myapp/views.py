# views.py
import os

import openpyxl
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.template import TemplateDoesNotExist
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.contrib.auth.hashers import make_password, check_password

import pdfkit
from django.conf import settings
from django.template.loader import render_to_string
from django.core.files.base import ContentFile

from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from .models import (
    Company, Site, User, SiteAssignment, Material,
    Rental, RentalDetail, RentalMovement, RentalMovementDetail, AIRecognition, ReturnRequest, ReturnRequestDetail
)

from .serializer import (
    CompanySerializer, SiteSerializer, UserSerializer, UserCreateSerializer,
    MaterialSerializer, SiteAssignmentSerializer,
    RentalSerializer, RentalDetailSerializer,
    RentalMovementSerializer, RentalMovementDetailSerializer, AIRecognitionSerializer
)

# =========================================================
# 간단 인증 (JWT 없이)
# =========================================================
def get_request_user(request):
    user_id = request.headers.get("X-USER-ID")
    if not user_id:
        return None
    try:
        return User.objects.get(id=int(user_id))
    except (User.DoesNotExist, ValueError):
        return None

def is_office_manager(user):
    return user and user.role == "officeManager"

def is_office_worker(user):
    return user and user.role == "officeWorker"

def is_site_manager(user):
    return user and user.role == "siteManager"

def is_site_worker(user):
    return user and user.role == "siteWorker"

def is_superadmin(user):
    return user and user.role in ("superadmin", "superAdmin")

# =========================================================
# 회사
# =========================================================
class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer

# 현장
class SiteViewSet(viewsets.ModelViewSet):
    queryset = Site.objects.all()
    serializer_class = SiteSerializer

    def get_queryset(self):
        me = get_request_user(self.request)
        if not me:
            return Site.objects.none()

        #  타사와 대여 계약을 맺기 위해 사무실 관리자/직원은 모든 현장을 볼 수 있게 허용
        if is_superadmin(me) or is_office_manager(me) or is_office_worker(me):
            # 프론트엔드에서 특정 회사(예: IBK)의 현장만 보고 싶을 때 ?company_id=2 로 필터링도 지원
            company_id = self.request.query_params.get('company_id')
            if company_id:
                return Site.objects.filter(company_id=company_id)
            return Site.objects.all()

        # 현장 관리자나 현장 직원은 자기 소속 회사의 현장만 볼 수 있도록 제한
        return Site.objects.filter(company=me.company)

    # 현장 등록 권한 제약 (이 아래는 기존 코드 그대로 유지)
    def create(self, request, *args, **kwargs):
        me = get_request_user(request)

        # 권한 체크: 시스템 관리자, 사무실 총관리자, 현장 총관리자만 현장 등록 가능
        if not me or not (is_superadmin(me) or is_office_manager(me) or is_site_manager(me)):
            return Response({"error": "현장을 등록할 권한이 없습니다."}, status=403)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # 등록 시 로그인한 관리자의 소속 회사로 강제 지정
        serializer.save(company=me.company)

        return Response(serializer.data, status=status.HTTP_201_CREATED)


# =========================================================
# 회원
# =========================================================
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_queryset(self):
        me = get_request_user(self.request)
        if not me: return User.objects.none()
        if is_superadmin(me): return User.objects.all()
        return User.objects.filter(company=me.company)

    def create(self, request, *args, **kwargs):
        # 1. 프론트엔드에서 넘어온 데이터 1차 검증
        serializer = UserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        company_id = request.data.get("company")

        # 2. 가입하려는 회사에 '사무실 총책임자(officeManager)'가 있는지 DB 확인
        has_office_manager = User.objects.filter(company_id=company_id, role="officeManager").exists()

        # 3. 상황에 맞게 초기 권한 셋팅
        if has_office_manager:
            assigned_role = "officeWorker"
        else:
            assigned_role = "siteWorker"

        # 4. 프론트에서 어떤 role을 보내든 무시하고, 백엔드가 판별한 role로 강제 저장 (보안)
        user = serializer.save(role=assigned_role, is_approved=False)

        return Response(
            {"message": "회원가입 성공", "user": UserSerializer(user).data},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="approve")
    def approve_user(self, request, pk=None):
        me = get_request_user(request)
        target = self.get_object()

        # 1. 승인 권한 체크 (슈퍼어드민, 사무실 관리자, 현장 관리자 모두 접근 가능)
        if not me or not (is_superadmin(me) or is_office_manager(me) or is_site_manager(me)):
            return Response({"error": "관리자만 승인 가능합니다."}, status=403)

        # 2. 소속 회사 체크 (슈퍼어드민 제외)
        if not is_superadmin(me) and target.company != me.company:
            return Response({"error": "같은 회사 직원만 승인 가능합니다."}, status=403)

        # 3. 역할에 따른 승인 권한 분리
        if target.role in ["officeManager", "siteManager"]:
            # 총관리자급 승인은 오직 superadmin만 가능
            if not is_superadmin(me):
                return Response({"error": "총관리자 승인은 시스템 최고관리자(superadmin)만 가능합니다."}, status=403)

        elif target.role == "officeWorker":
            # 사무실 직원 승인은 사무실 총관리자(또는 superadmin)만 가능
            if not (is_superadmin(me) or is_office_manager(me)):
                return Response({"error": "사무실 직원 승인은 사무실 총관리자만 가능합니다."}, status=403)

        elif target.role == "siteWorker":
            # 현장 직원 승인은 현장 총관리자(또는 superadmin)만 가능
            if not (is_superadmin(me) or is_site_manager(me)):
                return Response({"error": "현장 직원 승인은 현장 총관리자만 가능합니다."}, status=403)

        target.is_approved = True
        target.save()

        return Response({"message": "승인 완료"})

    @action(detail=True, methods=["patch"], url_path="change-role")
    def change_role(self, request, pk=None):
        me = get_request_user(request)
        target_user = self.get_object()

        # 1. 변경 권한 체크 (현장 총관리자도 접근할 수 있도록 추가)
        if not me or not (is_superadmin(me) or is_office_manager(me) or is_site_manager(me)):
            return Response({"error": "권한이 없습니다."}, status=403)

        # 2. 소속 회사 체크 (슈퍼어드민 제외)
        if not is_superadmin(me) and target_user.company != me.company:
            return Response({"error": "타 회사 직원의 권한은 변경할 수 없습니다."}, status=403)

        new_role = request.data.get("role")

        if new_role not in ["officeManager", "officeWorker", "siteManager", "siteWorker"]:
            return Response({"error": "유효하지 않은 role입니다."}, status=400)

        # 시스템 최고 관리자 보호
        if target_user.role == "superadmin" or new_role == "superadmin":
            return Response({"error": "superadmin 권한은 건드릴 수 없습니다."}, status=400)

        # 3. 내 직급에 따라 바꿀 수 있는 권한 제한
        if not is_superadmin(me):
            # 총관리자(Manager)급으로 승격시키는 것은 superadmin만 가능하도록 방어
            if new_role in ["officeManager", "siteManager"]:
                return Response({"error": "총관리자 임명은 최고관리자만 가능합니다."}, status=403)

            # 사무실 총관리자는 '사무실 직원' 권한만 부여 가능
            if is_office_manager(me) and new_role != "officeWorker":
                return Response({"error": "사무실 총관리자는 사무실 직원(officeWorker) 권한만 부여할 수 있습니다."}, status=403)

            # 현장 총관리자는 '현장 직원' 권한만 부여 가능
            if is_site_manager(me) and new_role != "siteWorker":
                return Response({"error": "현장 총관리자는 현장 직원(siteWorker) 권한만 부여할 수 있습니다."}, status=403)

        target_user.role = new_role
        target_user.save()

        return Response({"message": "권한 변경 완료", "user": UserSerializer(target_user).data})

    @action(detail=True, methods=["patch"], url_path="assign-site")
    def assign_site(self, request, pk=None):
        """
        관리자가 직원을 특정 현장에 배정하거나 해제(null)함
        """
        me = get_request_user(request)
        target_user = self.get_object()
        site_id = request.data.get("site_id")

        # 1. 권한 체크
        if not me or not (is_superadmin(me) or is_office_manager(me) or is_site_manager(me)):
            return Response({"error": "권한이 없습니다."}, status=403)

        # 2. 소속 회사 체크
        if not is_superadmin(me) and target_user.company != me.company:
            return Response({"error": "타 회사 직원은 배정할 수 없습니다."}, status=403)

        # 3. 배정 대상 직급 체크 (수정된 부분!)
        # 최고관리자와 사무실 총책임자는 '프리패스'로 누구든 배정 가능!
        if not (is_superadmin(me) or is_office_manager(me)):
            # 현장 총책임자일 경우에만 '현장 직원'인지 깐깐하게 검사
            if is_site_manager(me) and target_user.role != "siteWorker":
                return Response({"error": "현장 총관리자는 현장 직원(siteWorker)만 배정할 수 있습니다."}, status=403)

        # 4. JavaScript의 가짜 빈 값들을 모두 차단!
        if site_id in [None, "", "null", "undefined", 0, "0"]:
            target_user.site = None
            target_user.save()
            return Response({
                "message": f"[{target_user.name}] 직원의 현장 배정이 해제되었습니다.",
                "user": UserSerializer(target_user).data
            })

        # 5. 현장 할당
        try:
            site = get_object_or_404(Site, id=int(site_id))
        except ValueError:
            return Response({"error": "잘못된 현장 ID 형식입니다."}, status=400)

        target_user.site = site
        target_user.save()

        return Response({
            "message": f"[{target_user.name}] 직원이 [{site.site_name}] 현장에 성공적으로 배정되었습니다.",
            "user": UserSerializer(target_user).data
        })

    @action(detail=True, methods=["patch"], url_path="update-info")
    def update_info(self, request, pk=None):
        """
        사용자 정보 수정 (본인)
        """
        me = get_request_user(request)
        user = self.get_object()

        if not me or me.id != user.id:
            return Response({"error": "본인만 수정 가능"}, status=403)

        if "password" in request.data:
            user.password = make_password(request.data["password"])
        if "name" in request.data:
            user.name = request.data["name"]
        if "phone" in request.data:
            user.phone = request.data["phone"]

        user.save()
        return Response({"message": "정보 수정 완료", "user": UserSerializer(user).data})

    @action(detail=False, methods=["post"], url_path="check-email")
    def check_email(self, request):
        """
        아이디(이메일) 중복 확인 API
        """
        email = request.data.get("email")

        if not email:
            return Response({"error": "이메일을 입력해주세요."}, status=400)

        # User 테이블에 해당 이메일이 이미 존재하는지 확인
        is_duplicate = User.objects.filter(email=email).exists()

        if is_duplicate:
            return Response({
                "is_duplicate": True,
                "message": "이미 사용 중인 아이디(이메일)입니다."
            }, status=200)
        else:
            return Response({
                "is_duplicate": False,
                "message": "사용 가능한 아이디(이메일)입니다."
            }, status=200)

class LoginView(APIView):
    """
    단순 로그인
    """
    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response({"error": "email/password 필요"}, status=400)

        user = User.objects.filter(email=email).first()
        if not user or not check_password(password, user.password):
            return Response({"error": "로그인 실패"}, status=401)

        if not user.is_approved:
            return Response({"error": "승인되지 않은 사용자"}, status=403)

        return Response({
            "message": "login_success",
            "user_id": user.id,
            "role": user.role,
            "company_id": user.company_id
        })


# =========================================================
# 자재
# =========================================================
class MaterialViewSet(viewsets.ModelViewSet):
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def create(self, request, *args, **kwargs):
        me = get_request_user(request)
        if not me or not (is_office_manager(me) or is_office_worker(me)):
            return Response({"error": "권한 없음"}, status=403)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        serializer.save(company=me.company)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["patch"], url_path="update-info")
    def update_info(self, request, pk=None):
        """
        자재 정보(수량, 단가, 로스, 수리 등) 통합 수정
        """
        me = get_request_user(request)
        # 권한 체크: 사무실 관리자/직원만 가능
        if not me or not (is_office_manager(me) or is_office_worker(me)):
            return Response({"error": "권한 없음"}, status=403)

        material = self.get_object()


        if "price" in request.data:
            material.price = request.data["price"]
        if "total_qty" in request.data:
            material.total_qty = request.data["total_qty"]
        if "loss_qty" in request.data:
            material.loss_qty = request.data["loss_qty"]
        if "repaired_qty" in request.data:
            material.repaired_qty = request.data["repaired_qty"]

        # 2. DB 저장 전 유효성 검사 (숫자 형식 등)
        try:
            material.save()
        except Exception as e:
            return Response({"error": f"저장 중 오류 발생: {str(e)}"}, status=400)

        return Response({
            "message": "자재 정보가 성공적으로 수정되었습니다.",
            "user": me.name,  # 누가 수정했는지 확인용
            "data": MaterialSerializer(material).data
        })

# =========================================================
# 1. Rental (계약 총괄 관리 - 수정됨)
# =========================================================
class RentalViewSet(viewsets.ModelViewSet):
    queryset = Rental.objects.all()
    serializer_class = RentalSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        me = get_request_user(self.request)
        if not me: return Rental.objects.none()
        if is_superadmin(me): return Rental.objects.all()

        if is_office_manager(me):
            return Rental.objects.filter(
                Q(provider_company=me.company) | Q(requester_company=me.company)
            )
        if is_office_worker(me):
            return Rental.objects.filter(requester_user=me)

        # 현장 총책임자와 현장 직원의 권한 분리
        if is_site_manager(me):
            # 현장 총책임자는 '우리 회사가 담당하는 모든 현장'의 계약을 조회 가능
            if me.company:
                return Rental.objects.filter(site__company=me.company)
            return Rental.objects.none()

        elif is_site_worker(me):
            # 현장 직원은 본인이 배정된 현장의 계약만 조회 가능
            if me.site:
                return Rental.objects.filter(site=me.site)
            return Rental.objects.none()

        return Rental.objects.none()

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """ [1단계] 견적서 업로드 → 파싱 결과만 반환 (DB 저장 안 함) """
        me = get_request_user(request)
        if not me: return Response({"error": "로그인 필요"}, status=401)

        excel_file = request.FILES.get("quotation_file")
        provider_company_id = request.data.get("provider_company")
        site_id = request.data.get("site")

        provider_company = get_object_or_404(Company, id=provider_company_id)
        site = get_object_or_404(Site, id=site_id)

        rental = Rental.objects.create(
            requester_user=me, requester_company=me.company,
            provider_company=provider_company, site=site,
            status="DRAFT",  # ← REQUESTED → DRAFT로 변경 (아직 확정 전)
            quotation_file=excel_file,
        )

        parsed_items = []  # DB 저장 대신 파싱 결과를 리스트로 반환

        if excel_file and excel_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            try:
                sheet = wb['비계견적서']
            except KeyError:
                sheet = wb['비계 견적서']

            SKIP_KEYWORDS = {'소계', '소                 계', '총 공 사 금 액', '단 수 정 리'}

            def is_skip_row(name: str) -> bool:
                name = name.strip()
                if name[0].isdigit() and '.' in name: return True
                if name.startswith('*') or name.startswith('`'): return True
                if any(kw in name for kw in SKIP_KEYWORDS): return True
                return False

            for row in range(11, 500):
                material_name = sheet.cell(row=row, column=2).value
                if not material_name: break
                material_name = str(material_name).strip()
                if is_skip_row(material_name): continue

                planned_qty = sheet.cell(row=row, column=5).value

                # DB 매칭 여부 확인 (저장은 안 함)
                material = Material.objects.filter(
                    company=provider_company,
                    material_name__icontains=material_name
                ).first()

                parsed_items.append({
                    "material_id": material.id if material else None,
                    "material_name": material_name,  # 견적서에서 읽은 이름
                    "matched_name": material.material_name if material else None,  # DB 매칭된 이름
                    "planned_qty": planned_qty or 0,
                    "is_matched": material is not None,  # 프론트에서 빨간 표시용
                })

        return Response({
            "rental_id": rental.id,
            "parsed_items": parsed_items,  # 프론트가 이걸 화면에 보여줌
        }, status=201)

    @action(detail=True, methods=["post"], url_path="confirm-details")
    def confirm_details(self, request, pk=None):
        """
        [2단계] 프론트에서 수정 완료 후 RentalDetail 확정 저장
        body: { "items": [{"material_id": 1, "planned_qty": 100}, ...] }
        """
        me = get_request_user(request)
        rental = self.get_object()

        if rental.status != "DRAFT":
            return Response({"error": "이미 확정된 계약입니다."}, status=400)

        items = request.data.get("items", [])
        if not items:
            return Response({"error": "항목이 없습니다."}, status=400)

        for item in items:
            material = get_object_or_404(Material, id=item["material_id"])
            RentalDetail.objects.create(
                rental=rental,
                material=material,
                planned_qty=item.get("planned_qty", 0)
            )

        rental.status = "REQUESTED"
        rental.save()

        return Response({"message": "견적 확정 완료. 승인 대기 중입니다."})

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        """ [2단계] 대여 승인 """
        me = get_request_user(request)
        rental = self.get_object()
        if not me or not is_office_manager(me): return Response({"error": "권한 없음"}, status=403)
        rental.status = "ACTIVE"
        rental.approved_by = me
        rental.approved_at = timezone.now()
        rental.save()
        return Response({"message": "계약 활성화 (ACTIVE) 완료. 입출고 회차를 생성할 수 있습니다."})

class RentalDetailViewSet(viewsets.ModelViewSet):
    queryset = RentalDetail.objects.all()
    serializer_class = RentalDetailSerializer

    def get_queryset(self):
        qs = RentalDetail.objects.all()
        rental_id = self.request.query_params.get('rental')
        if rental_id:
            qs = qs.filter(rental_id=rental_id)
        return qs

# =========================================================
# 2. RentalMovement (입출고 회차 및 서명/문서 관리 - 핵심 신규 추가)
# =========================================================
class RentalMovementViewSet(viewsets.ModelViewSet):
    queryset = RentalMovement.objects.all()
    serializer_class = RentalMovementSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        me = get_request_user(self.request)
        if not me: return RentalMovement.objects.none()
        if is_superadmin(me): return RentalMovement.objects.all()

        # office_manager와 office_worker 모두 '우리 회사'와 관련된 계약은 다 볼 수 있게 통합!
        if is_office_manager(me) or is_office_worker(me):
            return RentalMovement.objects.filter(
                Q(rental__provider_company=me.company) |
                Q(rental__requester_company=me.company)
            )

        if is_site_manager(me):
            if me.company:
                return RentalMovement.objects.filter(rental__site__company=me.company)
            return RentalMovement.objects.none()

        elif is_site_worker(me):
            if me.site:
                return RentalMovement.objects.filter(rental__site=me.site)
            return RentalMovement.objects.none()

        return RentalMovement.objects.none()

    @action(detail=False, methods=["post"], url_path="create-out")
    def create_out_movement(self, request):
        me = get_request_user(request)
        rental_id = request.data.get("rental_id")
        rental = get_object_or_404(Rental, id=rental_id)

        movement = RentalMovement.objects.create(rental=rental, movement_type="OUT", status="OUT_CREATED",
                                                 created_by=me)

        import json
        details_data = json.loads(request.data.get("details", "[]"))
        for d in details_data:
            rd = get_object_or_404(RentalDetail, id=d["rental_detail_id"])
            RentalMovementDetail.objects.create(movement=movement, rental_detail=rd, material=rd.material,
                                                request_qty=d.get("request_qty", 0))
        return Response(RentalMovementSerializer(movement).data, status=201)

    @action(detail=True, methods=["post"], url_path="office-out-sign")
    def office_out_sign(self, request, pk=None):
        me = get_request_user(request)
        movement = self.get_object()

        movement.office_out_photo = request.FILES.get("photo")
        movement.office_out_signature = request.FILES.get("signature")
        movement.office_out_signed_by = me
        movement.status = "OUT_IN_TRANSIT"
        movement.save()

        import json
        details_data = json.loads(request.data.get("details", "[]"))
        for d in details_data:
            md = get_object_or_404(RentalMovementDetail, id=d["movement_detail_id"])
            md.office_out_qty = d.get("office_out_qty", md.request_qty)
            md.save()
            md.material.total_qty -= md.office_out_qty
            md.material.save()

        return Response({"message": "사무실 출고 서명 완료 및 본사 재고 차감됨"})

    @action(detail=True, methods=["post"], url_path="site-receive-sign")
    def site_receive_sign(self, request, pk=None):
        me = get_request_user(request)
        movement = self.get_object()

        movement.site_receive_photo = request.FILES.get("photo")
        movement.site_receive_signature = request.FILES.get("signature")
        movement.site_receive_signed_by = me
        movement.status = "OUT_COMPLETED"
        movement.save()

        import json
        details_data = json.loads(request.data.get("details", "[]"))
        for d in details_data:
            md = get_object_or_404(RentalMovementDetail, id=d["movement_detail_id"])
            md.site_receive_qty = d.get("site_receive_qty", md.office_out_qty)
            md.out_diff_qty = md.office_out_qty - md.site_receive_qty
            md.save()

        return Response({"message": "현장 수령 확인 완료"})

    @action(detail=False, methods=["post"], url_path="create-return")
    def create_return_movement(self, request):
        me = get_request_user(request)
        rental_id = request.data.get("rental_id")
        rental = get_object_or_404(Rental, id=rental_id)

        movement = RentalMovement.objects.create(
            rental=rental, movement_type="RETURN",
            status="RETURN_CREATED", created_by=me
        )

        import json
        details_data = json.loads(request.data.get("details", "[]"))
        for d in details_data:
            rd = get_object_or_404(RentalDetail, id=d["rental_detail_id"])
            RentalMovementDetail.objects.create(
                movement=movement, rental_detail=rd, material=rd.material,
                site_request_return_qty=d.get("request_qty", 0)
            )

        # 반납 신청 ACCEPTED 처리
        return_request_id = request.data.get("return_request_id")
        if return_request_id:
            ReturnRequest.objects.filter(id=return_request_id).update(status="ACCEPTED")

        return Response(RentalMovementSerializer(movement).data, status=201)

    @action(detail=True, methods=["post"], url_path="site-return-sign")
    def site_return_sign(self, request, pk=None):
        me = get_request_user(request)
        movement = self.get_object()

        movement.site_return_photo = request.FILES.get("photo")
        movement.site_return_signature = request.FILES.get("signature")
        movement.site_return_signed_by = me
        movement.status = "RETURN_IN_TRANSIT"
        movement.save()

        import json
        details_data = json.loads(request.data.get("details", "[]"))
        for d in details_data:
            md = get_object_or_404(RentalMovementDetail, id=d["movement_detail_id"])
            md.return_qty = d.get("return_qty", md.site_request_return_qty)
            md.save()

        return Response({"message": "현장 반납 서명 완료"})

    @action(detail=True, methods=["post"], url_path="office-in-sign")
    def office_in_sign(self, request, pk=None):
        me = get_request_user(request)
        movement = self.get_object()

        movement.office_in_photo = request.FILES.get("photo")
        movement.office_in_signature = request.FILES.get("signature")
        movement.office_in_signed_by = me
        movement.status = "RETURN_COMPLETED"
        movement.save()

        import json
        details_data = json.loads(request.data.get("details", "[]"))

        for d in details_data:
            md = get_object_or_404(RentalMovementDetail, id=d["movement_detail_id"])
            md.final_in_qty = d.get("final_in_qty", 0)
            md.loss_qty = d.get("loss_qty", 0)
            md.broken_qty = d.get("broken_qty", 0)
            md.discarded_qty = d.get("discarded_qty", 0)
            md.return_diff_qty = md.return_qty - md.final_in_qty
            md.save()

            md.material.total_qty += md.final_in_qty
            md.material.repaired_qty += md.broken_qty
            md.material.loss_qty += md.loss_qty
            md.material.disposed_qty += md.discarded_qty
            md.material.save()

        rental = movement.rental
        all_returned = all(rd.is_fully_returned for rd in rental.details.all())
        if all_returned:
            rental.status = "COMPLETED"
            rental.save()

        return Response({"message": "공장 입고 검수 완료 및 재고 복구됨"})

    @action(detail=True, methods=["get"], url_path="download-dispatch")
    def download_dispatch(self, request, pk=None):
        movement = self.get_object()
        if movement.movement_type != "OUT":
            return Response({"error": "출고 회차가 아닙니다."}, status=400)

        template_path = os.path.join(settings.BASE_DIR, 'myapp', 'templates', 'excel_templates', '반출증 양식.xlsx')
        wb = openpyxl.load_workbook(template_path)

        details = list(movement.movement_details.all())

        # 자재명 기준으로 시트 분기
        # material.category 필드가 있으면 그걸 쓰고,
        # 없으면 자재명에 '비계' 포함 여부로 판단
        def get_sheet_type(detail):
            name = detail.material.material_name or ""
            category = getattr(detail.material, 'category', '') or ""
            if '비계' in name or '비계' in category:
                return '비계'
            return '동바리'

        # 동바리/비계로 분리
        dongbari_details = [d for d in details if get_sheet_type(d) == '동바리']
        bigye_details = [d for d in details if get_sheet_type(d) == '비계']

        # 동바리 시트 채우기
        if dongbari_details:
            sheet = wb['동바리']
            sheet.cell(row=2, column=5).value = movement.rental.requester_company.company_name  # 거래처: D열=4
            sheet.cell(row=3, column=5).value = movement.rental.site.site_name  # 현장명
            sheet.cell(row=4, column=5).value = timezone.now().strftime('%Y-%m-%d')  # 날짜

            for i, detail in enumerate(dongbari_details):
                row = 8 + i
                sheet.cell(row=row, column=2).value = detail.material.material_name  # 자재명
                sheet.cell(row=row, column=3).value = detail.material.spec  # 규격
                sheet.cell(row=row, column=7).value = detail.office_out_qty or detail.request_qty or 0  # 임대 개수

        # 비계 시트 채우기
        if bigye_details:
            sheet = wb['비계']
            sheet.cell(row=2, column=6).value = movement.rental.requester_company.company_name  # 거래처: E열=5
            sheet.cell(row=3, column=6).value = movement.rental.site.site_name
            sheet.cell(row=4, column=6).value = timezone.now().strftime('%Y-%m-%d')

            for i, detail in enumerate(bigye_details):
                row = 8 + i
                sheet.cell(row=row, column=2).value = detail.material.material_name  # 품명
                sheet.cell(row=row, column=4).value = detail.material.spec  # 규격: D열=4
                sheet.cell(row=row, column=8).value = detail.office_out_qty or detail.request_qty or 0  # 임대 개수: H열=8

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename=Rental{movement.rental.id}_OUT{movement.id}_dispatch.xlsx'
        wb.save(response)
        return response

    @action(detail=True, methods=["get"], url_path="download-invoice")
    def download_invoice(self, request, pk=None):
        movement = self.get_object()
        if movement.movement_type != "OUT": return Response({"error": "출고 회차가 아닙니다."}, status=400)

        template_path = os.path.join(settings.BASE_DIR, 'myapp', 'templates', 'excel_templates', '출고송장.xlsx')
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active

        sheet['F1'] = movement.rental.requester_company.company_name
        sheet['F2'] = movement.rental.site.site_name
        sheet['F6'] = movement.rental.requester_user.name

        for i, detail in enumerate(movement.movement_details.all()):
            row = 9 + i
            sheet.cell(row=row, column=2).value = detail.material.material_name
            sheet.cell(row=row, column=4).value = detail.material.spec
            sheet.cell(row=row, column=7).value = detail.office_out_qty or 0

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename=Rental{movement.rental.id}_OUT{movement.id}_invoice.xlsx'
        wb.save(response)
        return response

    @action(detail=True, methods=["get"], url_path="download-return-confirm")
    def download_return_confirm(self, request, pk=None):
        movement = self.get_object()
        if movement.movement_type != "RETURN": return Response({"error": "반납 회차가 아닙니다."}, status=400)

        template_path = os.path.join(settings.BASE_DIR, 'myapp', 'templates', 'excel_templates', '공장_입고검수확인서.xlsx')
        wb = openpyxl.load_workbook(template_path)
        sheet = wb['잔량']

        sheet['C4'] = movement.rental.site.site_name
        sheet['C2'] = movement.created_at.strftime('%Y-%m-%d')

        for i, detail in enumerate(movement.movement_details.all()):
            row = 12 + i
            sheet.cell(row=row, column=2).value = detail.material.material_name
            sheet.cell(row=row, column=5).value = detail.material.spec
            sheet.cell(row=row, column=8).value = detail.final_in_qty or 0
            sheet.cell(row=row, column=9).value = detail.loss_qty or 0

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename=Rental{movement.rental.id}_RETURN{movement.id}_return_confirm.xlsx'
        wb.save(response)
        return response


class AIRecognitionViewSet(viewsets.ModelViewSet):
    queryset = AIRecognition.objects.all()
    serializer_class = AIRecognitionSerializer

    def create(self, request, *args, **kwargs):
        me = get_request_user(request)

        # 인증된 모든 직원(사무실 관리자/직원, 현장 관리자/직원)이 결과 저장 가능
        if not me or not (is_superadmin(me) or is_office_manager(me) or is_office_worker(me) or is_site_manager(
                me) or is_site_worker(me)):
            return Response({"error": "AI 판독 결과를 저장할 권한이 없습니다."}, status=403)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        ai = serializer.save()

        # movement_detail에 연결된 경우
        if ai.movement_detail and ai.detected_qty is not None:
            ai.movement_detail.final_in_qty = ai.detected_qty
            ai.movement_detail.save()

        # return_request_detail에 연결된 경우 추가
        if ai.return_request_detail and ai.detected_qty is not None:
            ai.return_request_detail.request_qty = ai.detected_qty
            ai.return_request_detail.save()

        return Response(serializer.data, status=201)

class ReturnRequestViewSet(viewsets.ModelViewSet):
    queryset = ReturnRequest.objects.all()
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        me = get_request_user(self.request)
        if not me: return ReturnRequest.objects.none()
        if is_superadmin(me) or is_office_manager(me):
            return ReturnRequest.objects.filter(
                rental__provider_company=me.company
            ).order_by('-created_at')

        if is_site_manager(me):
            if me.company:
                return ReturnRequest.objects.filter(rental__site__company=me.company)
            return ReturnRequest.objects.none()

        elif is_site_worker(me):
            if me.site:
                return ReturnRequest.objects.filter(rental__site=me.site)
            return ReturnRequest.objects.none()

        return ReturnRequest.objects.none()

    def create(self, request, *args, **kwargs):
        """ 현장 총책임자가 반납 신청 제출 """
        me = get_request_user(request)
        if not me or not (is_site_manager(me) or is_superadmin(me)):
            return Response({"error": "현장 총책임자만 반납 신청 가능합니다."}, status=403)

        rental = get_object_or_404(Rental, id=request.data.get("rental_id"))

        return_request = ReturnRequest.objects.create(
            rental=rental,
            created_by=me,
            photo=request.FILES.get("photo"),
            signature=request.FILES.get("signature"),
        )

        import json
        details_data = json.loads(request.data.get("details", "[]"))
        for d in details_data:
            rd = get_object_or_404(RentalDetail, id=d["rental_detail_id"])
            ReturnRequestDetail.objects.create(
                return_request=return_request,
                rental_detail=rd,
                request_qty=d.get("request_qty", 0)
            )

        return Response(self._serialize(return_request), status=201)

    def list(self, request, *args, **kwargs):
        """ 사무실 총책임자가 PENDING 목록 조회 """
        qs = self.get_queryset().filter(status="PENDING")
        return Response([self._serialize(rr) for rr in qs])

    def _serialize(self, rr):
        return {
            "id": rr.id,
            "rental": rr.rental.id,
            "site_name": rr.rental.site.site_name,
            "status": rr.status,
            "created_at": rr.created_at,
            "details": [
                {
                    "rental_detail_id": d.rental_detail.id,
                    "material_name": d.rental_detail.material.material_name,
                    "spec": d.rental_detail.material.spec,
                    "request_qty": d.request_qty,
                }
                for d in rr.details.all()
            ]
        }

#==============
import requests

class DetectAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if "image" not in request.FILES:
            return Response({"error": "image required"}, status=400)

        image = request.FILES["image"]

        # 파일 포인터를 처음으로 되돌린 후 읽기
        image.seek(0)
        files = {
            "image": (image.name, image.read(), image.content_type)
        }

        try:
            ai_response = requests.post(
                "http://172.17.0.1:8001/detect",
                files=files,
                timeout=30
            )
            return Response(ai_response.json(), status=ai_response.status_code)

        except Exception as e:
            return Response(
                {
                    "error": "AI 서버 통신 중 오류가 발생했습니다.",
                    "detail": str(e)
                },
                status=502
            )
